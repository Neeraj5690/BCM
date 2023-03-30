import datetime
import sys
if "C:/Users/Neeraj/PycharmProjects/BCM" not in sys.path:
    sys.path.append("C:/Users/Neeraj/PycharmProjects/BCM")
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.common.by import By
from chrome.LatestChrome import ChromeCls
from TestEnvironment.GlobalClassMethods.MasterDataExcelReader import DataReadMaster
from TestEnvironment.GlobalElementAction.ElementAction import ElementActionCls
from TestEnvironment.GlobalElementAction.SafeToElementAction import SafeToElementActionCls, SafeToVerify
from TestEnvironment.GlobalElementPresent.ElementPresent import ElementPresentCls
from TestEnvironment.GlobalLoader.Loader import LoaderCls

MdataSheetTab = "test_Smoke_Admin_Manager"
PageName = "Admin"
@allure.step("Entering username ")
def enter_username(username):
    driver.find_element(By.ID, "un").send_keys(username)


@allure.step("Entering password ")
def enter_password(password):
    driver.find_element(By.ID, "pw").send_keys(password)


@pytest.fixture()
def test_setup():
    sys.path.append("/chrome")
    print(ChromeCls.NewChromePathChrCls)

    global driver, TestResult, TestResultStatus, path, FundNameList, FundNameListAfterRemove, ct, Exe, D1, D2, d1, d2, DollarDate

    TestResult = []
    TestResultStatus = []
    TestFailStatus = []
    FailStatus = "Pass"
    Exe = "Yes"

    path = DataReadMaster.Path + DataReadMaster.GlobalData(MdataSheetTab,
                                                           "ParentDirectory") + DataReadMaster.GlobalData(
        MdataSheetTab, "Directory") + DataReadMaster.GlobalData(MdataSheetTab, "SubDirectory")
    FundNameList = []
    FundNameListAfterRemove = []

    ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
    ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

    today = datetime.date.today()
    D1 = today.strftime("%Y-%m-%d")
    d1 = D1
    DollarDate = datetime.datetime.strptime(d1, '%Y-%m-%d')
    DollarDate = "$" + DollarDate.date().__str__() + "$"
    d1 = datetime.datetime.strptime(D1, "%Y-%m-%d")

    Exe = DataReadMaster.GlobalData(MdataSheetTab, "Execution")

    # --------Login to the application-----------------------
    if Exe == "Yes":
        ChromeCls.ChromeMeth()
        driver = webdriver.Chrome(executable_path=ChromeCls.NewChromePath1ChrCls)
        driver.implicitly_wait(10)
        driver.maximize_window()
        driver.get(DataReadMaster.GlobalData("GlobalData", "URLManager"))
        enter_username(DataReadMaster.GlobalData("GlobalData", "ManagerUsername"))
        enter_password(DataReadMaster.GlobalData("GlobalData", "ManagerPassword"))
        driver.find_element(By.XPATH, DataReadMaster.GlobalData("GlobalData", "LoginSubmit")).click()

    yield
    if Exe == "Yes":
        class PDF(FPDF):
            def header(self):
                self.image(path + 'EmailReportContent/Logo.png', 10, 8, 33)
                self.set_font('Arial', 'B', 15)
                self.cell(73)
                self.set_text_color(0, 0, 0)
                self.cell(35, 10, ' Test Report ', 1, 1, 'B')
                self.set_font('Arial', 'I', 10)
                self.cell(150)
                self.cell(30, 10, ctReportHeader, 0, 0, 'C')
                self.ln(20)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(0, 0, 0)
                self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font('Times', '', 12)
        pdf.cell(0, 10, "Test Case Name:  " + DataReadMaster.GlobalData(MdataSheetTab, "PDFTestName"), 0, 1)
        pdf.multi_cell(0, 10,
                       "Description:  " + DataReadMaster.GlobalData(MdataSheetTab, "PDFDescription"), 0, 1)
        for i1 in range(len(TestResult)):
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            if (TestResultStatus[i1] == "Fail"):
                # print("Fill Red color")
                pdf.set_text_color(255, 0, 0)
                TestFailStatus.append("Fail")
            TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
            pdf.multi_cell(0, 7, str(i1 + 1) + ")  " + TestName1, 0, 1, fill=True)
            TestFailStatus.append("Pass")
        pdf.output(DataReadMaster.GlobalData(MdataSheetTab, "TestName") + "_" + ct + ".pdf")

        # -----------To check if any failed Test case present-------------------
        for io in range(len(TestResult)):
            if TestFailStatus[io] == "Fail":
                FailStatus = "Fail"
        # ---------------------------------------------------------------------

        # -----------To add test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        print()
        check = DataReadMaster.GlobalData(MdataSheetTab, "TestName")
        PdfName = DataReadMaster.GlobalData(MdataSheetTab, "TestName") + "_" + ct + ".pdf"
        checkcount = 0

        for i in range(1, 100):
            if sheet.cell(i, 1).value == None:
                if checkcount == 0:
                    sheet.cell(row=i, column=1).value = check
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData(MdataSheetTab,
                                                                                  "TestDirectoryName")
                    sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData(MdataSheetTab,
                                                                                  "PDFDescription")
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
                wb.save(loc)
                break
            else:
                if sheet.cell(i, 1).value == check:
                    if checkcount == 0:
                        sheet.cell(row=i, column=2).value = PdfName
                        sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData(MdataSheetTab,
                                                                                      "TestDirectoryName")
                        sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData(MdataSheetTab,
                                                                                      "PDFDescription")
                        sheet.cell(row=i, column=5).value = FailStatus
                        checkcount = 1
        # -----------------------------------------------------------------------------
        driver.quit()


@pytest.mark.smoke
def test_AllModules(test_setup):
    if Exe == "Yes":
        try:
            # ---------------------------Verify Plans page-----------------------------
            driver.find_element(By.XPATH,
                                DataReadMaster.GlobalData(MdataSheetTab, "HomePage")).click()
            PageTitleExpected = "Admin - Demo - Benefits Coverage Manager"
            LoaderCls.LoaderMeth(driver)
            try:
                PageTitleFound = driver.title
                assert PageTitleFound in PageTitleExpected, PageName + " module was not able to open"
                TestResult.append(PageName + " module was able to open successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " module was not able to open")
                TestResultStatus.append("Fail")

            # --------------------------Top Big Buttons-------------------------
            # -----Create or Edit Plan Cost Share Validations Button------------
            ElementVerify = "Create or Edit Plan Cost Share Validations tile"
            ElementExpected = "Manage Validation for Plan Cost Shares"
            MdataSheetItem = "CreateEditPlanCostButton"
            MdataSheetItem2 = "CreateEditPlanCostButtonText"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            driver.find_element(By.XPATH,DataReadMaster.GlobalData(MdataSheetTab, "HomePage")).click()

            # -----Create or Edit Plan Service Button------------
            ElementVerify = "Create or Edit Plan Service tile"
            ElementExpected = "Service Name"
            MdataSheetItem = "CreateEditPlanServiceButton"
            MdataSheetItem2 = "CreateEditPlanServiceButtonText"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            driver.find_element(By.XPATH, DataReadMaster.GlobalData(MdataSheetTab, "HomePage")).click()

            # -----Edit Organizational Profile Button------------
            ElementVerify = "Edit Organizational Profile tile"
            ElementExpected = "Update Organization Profile"
            MdataSheetItem = "EditOrganizationalProfileButton"
            MdataSheetItem2 = "EditOrganizationalProfileButtonText"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            driver.find_element(By.XPATH, DataReadMaster.GlobalData(MdataSheetTab, "HomePage")).click()

        except Exception as Mainerror:
            print(Mainerror)
            stringMainerror = repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")
    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = DataReadMaster.GlobalData(MdataSheetTab, "TestName")

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------